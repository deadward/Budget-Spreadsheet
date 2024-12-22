from datetime import date, timedelta, datetime
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, numbers, PatternFill

class CalendarFormater:
    def __init__(self, sheet):
        self.sheet = sheet

    def apply_header_fromats(self, calendar_start_row):
        for col in range(1, 7):
            cell = self.sheet[f"{get_column_letter(col)}{calendar_start_row}"]
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def create_cal_headers(self, calendar_start_row):
        self.sheet[f"A{calendar_start_row}"] = "Date"
        self.sheet[f"B{calendar_start_row}"] = "Transaction Name"
        self.sheet[f"C{calendar_start_row}"] = "Amount"
        self.sheet[f"D{calendar_start_row}"] = "Checking Total"
        self.sheet[f"E{calendar_start_row}"] = "Additional Amount"
        self.sheet[f"F{calendar_start_row}"] = "Grand Total"

    def add_pending_row(self, calendar_start_row):
        pending_row = calendar_start_row + 1
        self.sheet[f"A{pending_row}"] = "Pending"
        self.sheet[f"A{pending_row}"].alignment = Alignment(horizontal="left")
        self.sheet[f"D{pending_row}"] = f"=B3+C23"
        self.sheet[f"D{pending_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        self.sheet[f"F{pending_row}"] = f"=D23+B4+B5+B6+B7"
        self.sheet[f"F{pending_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    
    # comment
    def create_calendar(self, calendar_start_row):
        self.create_cal_headers(calendar_start_row)
        self.apply_header_fromats(calendar_start_row)
        self.add_pending_row(calendar_start_row)

        # Generate the calendar year of dates
        start_date = date(2025, 1, 1)
        end_date = date(2026, 12, 31)
        current_row = calendar_start_row + 2
        current_date = start_date

        #pay day = january 14
        pay_start = 14
        start_adding_pay = False
        payday = 0

        while current_date <= end_date:
            # Default pay added indicator at beginning of loop
            pay_added_ind = False
            waste_added_ind = False

            # Write the date in the first column
            if current_date.day == 1:
                self.sheet[f"A{current_row}"] = current_date.strftime("%B %d")  # Month and day for the 1st of each month
            else:
                self.sheet[f"A{current_row}"] = current_date.day  # Day of the month

            # Aligning cell data
            self.sheet[f"A{current_row}"].alignment = Alignment(horizontal="right")

            # Check for first paycheck
            if pay_start == current_date.day:
                start_adding_pay = True

            # Check if it is payday
            if start_adding_pay is True:
                if payday == 0:
                    # Add the pay to the row
                    self.sheet[f"B{current_row}"] = "Pay"
                    self.sheet[f"C{current_row}"] = f"=E3"
                    self.sheet[f"C{current_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                    self.sheet[f"E{current_row}"] = f"=E4+E5+E6"
                    self.sheet[f"E{current_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

                    # Add formulas and formatting for Amount and Additional Amount
                    self.apply_formulas(self.sheet, current_row)

                    # Adding pay indicator so that bills won't overwrite it
                    pay_added_ind = True

                    # Incrementing payday
                    payday += 1
                elif payday == 13:
                    # Reseting payday
                    payday = 0
                else:
                    # Incrementing payday
                    payday += 1

            # Adding an absolutly moronoic section for Frontier Waste that is every 3 months
            frontier_dates = [
                datetime(2025, 3, 5).date(),
                datetime(2025, 6, 5).date(),
                datetime(2025, 9, 5).date(),
                datetime(2025, 12, 5).date(),
                datetime(2026, 3, 5).date(),
                datetime(2026, 6, 5).date(),
                datetime(2026, 9, 5).date(),
                datetime(2026, 12, 5).date(),
            ]

            ## mmove all of these if statements to their own individual functions
            if current_date in frontier_dates:
                # Increment a row if we already added Pay
                if pay_added_ind is True:
                    current_date += 1
                    # Setting pay indicator to False since we already incrememted passed it
                    pay_added_ind = False

                # Add the waste to the row
                self.sheet[f"B{current_row}"] = "Frontier Waste"
                self.sheet[f"C{current_row}"] = f"=B20"
                self.sheet[f"C{current_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                waste_added_ind = True

                # Add formulas and formatting for Amount and Additional Amount
                self.apply_formulas(self.sheet, current_row)

            # Check for bills due on this date
            bills_to_add = self.check_bill_date(current_date.day)

            if bills_to_add:
                # Incrementing a row to not overwrite pay
                if pay_added_ind is True:
                    current_row += 1

                if waste_added_ind is True:
                    current_row += 1

                for bill in bills_to_add:
                    # Add the bill name in column B
                    self.sheet[f"B{current_row}"] = bill

                    # Dynamically reference the amount from the Income table
                    self.sheet[f"C{current_row}"] = f"=B{self.bill_column(bill)}*-1"
                    self.sheet[f"C{current_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

                    # Add formulas and formatting for Amount and Additional Amount
                    self.apply_formulas(self.sheet, current_row)

                    # Move to the next row for additional bills on the same day
                    current_row += 1
            else:
                # Add formulas and formatting for Amount and Additional Amount
                self.apply_formulas(self.sheet, current_row)

                # No bills due; skip to the next row
                current_row += 1

            # Move to the next date
            current_date += timedelta(days=1)

        self.apply_conditional_formatting1(self.sheet, col="D", start_row=23, end_row=current_row-1)
        self.apply_conditional_formatting2(self.sheet, col="F", start_row=23, end_row=current_row-1)

        return self.sheet


    def check_bill_date(day):
        bill_due_dates = {
            "Mortgage": 1,
            "Student Loan": 27,
            "Auto Insurance": 29,
            "Verizon": 13,
            "FOR Water": 25,
            "Hulu": 28,
            "CPS": 12,
            "Spectrum": 18,
            "USAA": 12,
            "Citi": 2,
            "Chase": 1
        }

        # Return a list of bills due on the given day
        return [bill for bill, due_day in bill_due_dates.items() if due_day == day]


    def bill_column(self, bill):
        bill_column_mapping = {
            "Mortgage": 12,
            "Student Loan": 13,
            "Auto Insurance": 14,
            "Verizon": 15,
            "FOR Water": 16,
            "Hulu": 17,
            "CPS": 18,
            "Spectrum": 19
        }
        return bill_column_mapping.get(bill, 1)  # Default to column 1 if not found


    def apply_formulas(self, current_row):
        # Add formulas and formatting for Amount and Additional Amount
        self.sheet[f"D{current_row}"] = f"=D{current_row - 1}+C{current_row}"
        self.sheet[f"D{current_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        self.sheet[f"F{current_row}"] = f"=C{current_row}+E{current_row}+F{current_row - 1}"
        self.sheet[f"F{current_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE


    def apply_conditional_formatting1(self, col, start_row, end_row):
        # Define the column range (e.g., B2:B10)
        cell_range = f"{col}{start_row}:{col}{end_row}"

        # Define color fills
        # these can be constants
        red_fill = PatternFill(start_color="FF0D0D", end_color="FF0D0D", fill_type="solid")
        burn_orange_fill = PatternFill(start_color="FF4E11", end_color="FF4E11", fill_type="solid")
        orange_fill = PatternFill(start_color="FF8E15", end_color="FF8E15", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        green_fill = PatternFill(start_color="69B34C", end_color="69B34C", fill_type="solid")
        neon_green_fill = PatternFill(start_color="39FF14", end_color="39FF14", fill_type="solid")
        
        red_rule = FormulaRule(
            formula=[f"{col}{start_row}<499"],
            stopIfTrue=True,
            fill=red_fill
        )

        burn_orange_rule = FormulaRule(
            formula=[f"AND({col}{start_row}>=500, {col}{start_row}<=999)"],
            stopIfTrue=True,
            fill=burn_orange_fill
        )

        orange_rule = FormulaRule(
            formula=[f"AND({col}{start_row}>=1000, {col}{start_row}<=1499)"],
            stopIfTrue=True,
            fill=orange_fill
        )

        yellow_rule = FormulaRule(
            formula=[f"AND({col}{start_row}>=1500, {col}{start_row}<=1999)"],
            stopIfTrue=True,
            fill=yellow_fill
        )

        light_green_rule = FormulaRule(
            formula=[f"AND({col}{start_row}>=2000, {col}{start_row}<=3999)"],
            stopIfTrue=True,
            fill=light_green_fill
        )

        green_rule = FormulaRule(
            formula=[f"AND({col}{start_row}>=4000, {col}{start_row}<=5999)"],
            stopIfTrue=True,
            fill=green_fill
        )

        neon_green_rule = FormulaRule(
            formula=[f"{col}{start_row}>6000"],
            stopIfTrue=True,
            fill=neon_green_fill
        )

        # Apply rules to the sheet
        self.sheet.conditional_formatting.add(cell_range, red_rule)
        self.sheet.conditional_formatting.add(cell_range, burn_orange_rule)
        self.sheet.conditional_formatting.add(cell_range, orange_rule)
        self.sheet.conditional_formatting.add(cell_range, yellow_rule)
        self.sheet.conditional_formatting.add(cell_range, light_green_rule)
        self.sheet.conditional_formatting.add(cell_range, green_rule)
        self.sheet.conditional_formatting.add(cell_range, neon_green_rule)

    def apply_conditional_formatting2(self, col, start_row, end_row):
        # Define the column range (e.g., B2:B10)
        cell_range = f"{col}{start_row}:{col}{end_row}"

        # Define color fills
        red_fill = PatternFill(start_color="FF0D0D", end_color="FF0D0D", fill_type="solid")
        burn_orange_fill = PatternFill(start_color="FF4E11", end_color="FF4E11", fill_type="solid")
        orange_fill = PatternFill(start_color="FF8E15", end_color="FF8E15", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        green_fill = PatternFill(start_color="69B34C", end_color="69B34C", fill_type="solid")
        neon_green_fill = PatternFill(start_color="39FF14", end_color="39FF14", fill_type="solid")

        red_rule = FormulaRule(
            formula=[f"{col}{start_row}<9999"],
            stopIfTrue=True,
            fill=red_fill
        )

        burn_orange_rule = FormulaRule(
            formula=[f"AND({col}{start_row}>=10000, {col}{start_row}<=14999)"],
            stopIfTrue=True,
            fill=burn_orange_fill
        )

        orange_rule = FormulaRule(
            formula=[f"AND({col}{start_row}>=15000, {col}{start_row}<=19999)"],
            stopIfTrue=True,
            fill=orange_fill
        )

        yellow_rule = FormulaRule(
            formula=[f"AND({col}{start_row}>=20000, {col}{start_row}<=24999)"],
            stopIfTrue=True,
            fill=yellow_fill
        )

        light_green_rule = FormulaRule(
            formula=[f"AND({col}{start_row}>=25000, {col}{start_row}<=39999)"],
            stopIfTrue=True,
            fill=light_green_fill
        )

        green_rule = FormulaRule(
            formula=[f"AND({col}{start_row}>=40000, {col}{start_row}<=59999)"],
            stopIfTrue=True,
            fill=green_fill
        )

        neon_green_rule = FormulaRule(
            formula=[f"{col}{start_row}>60000"],
            stopIfTrue=True,
            fill=neon_green_fill
        )

        # Apply rules to the sheet
        self.sheet.conditional_formatting.add(cell_range, red_rule)
        self.sheet.conditional_formatting.add(cell_range, burn_orange_rule)
        self.sheet.conditional_formatting.add(cell_range, orange_rule)
        self.sheet.conditional_formatting.add(cell_range, yellow_rule)
        self.sheet.conditional_formatting.add(cell_range, light_green_rule)
        self.sheet.conditional_formatting.add(cell_range, green_rule)
        self.sheet.conditional_formatting.add(cell_range, neon_green_rule)