import openpyxl
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter

def format_table_header(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def create_budget_tables(sheet):
    # Account Totals
    sheet["A1"] = "Account Totals"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = "Account Name"
    sheet["B2"] = "Amount"
    format_table_header(sheet["A2"])
    format_table_header(sheet["B2"])

    accounts = ["Checking", "Savings", "Emergency", "CG Wax Studio", "Investment Checking"]
    amounts = [5000, 7000, 7500, 13000, 1000]
    for idx, (account, amount) in enumerate(zip(accounts, amounts), start=3):
        sheet[f"A{idx}"] = account
        sheet[f"B{idx}"] = amount
        sheet[f"B{idx}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Total row
    total_row = len(accounts) + 3
    sheet[f"A{total_row}"] = "Total Current"
    sheet[f"B{total_row}"] = f"=SUM(B3:B{total_row - 1})"
    sheet[f"B{total_row}"].font = Font(bold=True)
    sheet[f"B{total_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    sheet[f"A{total_row}"].font = Font(bold=True)

    # Income Table
    data_start_col = 4
    sheet[f"{get_column_letter(data_start_col)}1"] = "Income"
    sheet[f"{get_column_letter(data_start_col)}1"].font = Font(bold=True, size=14)
    sheet[f"{get_column_letter(data_start_col)}2"] = "Name"
    sheet[f"{get_column_letter(data_start_col + 1)}2"] = "Amount"
    format_table_header(sheet[f"{get_column_letter(data_start_col)}2"])
    format_table_header(sheet[f"{get_column_letter(data_start_col + 1)}2"])

    income_items = ["Checking Paycheck", "Savings Paycheck", "Emergency Paycheck", "Citi Paycheck", "Total Paycheck", "Bonus"]
    income_amounts = [3700, 211, 211, 125, "=SUM(E3:E6)", "=ROUND((173000 * 1 * 0.2) * 0.72, 0)"]
    for idx, (item, amount) in enumerate(zip(income_items, income_amounts), start=3):
        sheet[f"{get_column_letter(data_start_col)}{idx}"] = item
        sheet[f"{get_column_letter(data_start_col + 1)}{idx}"] = amount
        if isinstance(amount, (int, float)) or (isinstance(amount, str) and amount.startswith("=")):
            sheet[f"{get_column_letter(data_start_col + 1)}{idx}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Bills Table
    bills_start_row = total_row + 2
    sheet[f"A{bills_start_row}"] = "Bills"
    sheet[f"A{bills_start_row}"].font = Font(bold=True, size=14)
    sheet[f"A{bills_start_row + 1}"] = "Bill Name"
    sheet[f"B{bills_start_row + 1}"] = "Bill Amount"
    format_table_header(sheet[f"A{bills_start_row + 1}"])
    format_table_header(sheet[f"B{bills_start_row + 1}"])

    bills = ["Mortgage", "Student Loan", "Auto Insurance", "Verizon", "FOR Water", "Hulu", "CPS", "Spectrum", "Frontier Waste"]
    bill_amounts = [4200, 432, 289, 192, 85, 103, 300, 105, 92]
    for idx, (bill, amount) in enumerate(zip(bills, bill_amounts), start=bills_start_row + 2):
        sheet[f"A{idx}"] = bill
        sheet[f"B{idx}"] = amount
        sheet[f"B{idx}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Debt Table
    credit_cards_start_row = bills_start_row
    sheet[f"D{credit_cards_start_row}"] = "Debt"
    sheet[f"D{credit_cards_start_row}"].font = Font(bold=True, size=14)
    sheet[f"D{credit_cards_start_row + 1}"] = "Card Name"
    sheet[f"E{credit_cards_start_row + 1}"] = "Amount"
    format_table_header(sheet[f"D{credit_cards_start_row + 1}"])
    format_table_header(sheet[f"E{credit_cards_start_row + 1}"])

    credit_cards = ["USAA", "Citi", "Chase", "Student Loan", "Subaru"]
    credit_card_amounts = [500, 0, 1200, 19400, 32850]
    for idx, (card, amount) in enumerate(zip(credit_cards, credit_card_amounts), start=credit_cards_start_row + 2):
        sheet[f"D{idx}"] = card
        sheet[f"E{idx}"] = amount
        sheet[f"E{idx}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    return sheet
