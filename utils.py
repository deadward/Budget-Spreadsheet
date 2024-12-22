from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def adjust_column_widths(sheet):
    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in sheet.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width
